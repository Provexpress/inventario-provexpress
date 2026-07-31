import os
import json
import sys
import openpyxl
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\especialista.prevent\OneDrive - PROVEXPRESS SAS\Excels Invetario\plantilla_inventario_matriz.xlsx"
PORT = 8000

class SyncHandler(BaseHTTPRequestHandler):
    def _set_headers(self, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_OPTIONS(self):
        self._set_headers(200)

    def do_GET(self):
        if self.path == '/api/status':
            self._set_headers(200)
            self.wfile.write(json.dumps({'status': 'online', 'excel_path': EXCEL_PATH}).encode('utf-8'))
        else:
            self._set_headers(404)
            self.wfile.write(json.dumps({'error': 'Not found'}).encode('utf-8'))

    def do_POST(self):
        if self.path == '/api/sync_and_send':
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            
            try:
                payload = json.loads(post_data.decode('utf-8'))
                products = payload.get('products', [])
                recipient = payload.get('recipient', 'especialista.preventa@provexpress.com.co')
                
                # 1. Update Excel Matrix
                sync_excel(products)
                
                # 2. Send Outlook Email
                send_outlook(products, recipient)
                
                self._set_headers(200)
                self.wfile.write(json.dumps({'success': True, 'message': 'Inventario sincronizado y correo enviado.'}).encode('utf-8'))
            except Exception as e:
                self._set_headers(500)
                self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode('utf-8'))

def sync_excel(products):
    if not os.path.exists(EXCEL_PATH):
        return

    # Try COM active Excel first
    excel_app = None
    wb_com = None
    try:
        import win32com.client as win32
        excel_app = win32.GetActiveObject("Excel.Application")
        for wb_item in excel_app.Workbooks:
            if "plantilla_inventario_matriz" in wb_item.Name.lower():
                wb_com = wb_item
                break
    except Exception:
        pass

    now = datetime.now()
    day_num = now.day
    month_names = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    month_label = f"{month_names[now.month-1]} {now.year}"

    if wb_com:
        ws_matriz = wb_com.Worksheets("Matriz Inventario")
        last_rm = ws_matriz.Cells(ws_matriz.Rows.Count, 4).End(-4162).Row
        row_map = {}
        for r in range(2, last_rm + 1):
            m_v = ws_matriz.Cells(r, 1).Value
            np_v = ws_matriz.Cells(r, 4).Value
            if m_v == month_label and np_v:
                row_map[str(np_v).strip()] = r

        for item in products:
            np_v = str(item.get('np', '')).strip()
            cant = int(item.get('cantidad_actual', 0))
            if np_v in row_map:
                target_row = row_map[np_v]
                ws_matriz.Cells(target_row, 10).Value = cant
                day_col_idx = 10 + day_num
                ws_matriz.Cells(target_row, day_col_idx).Value = cant
        wb_com.Save()
    else:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
        ws = wb["Matriz Inventario"]
        row_map = {}
        for r in range(2, ws.max_row + 1):
            m_v = ws.cell(r, 1).value
            np_v = ws.cell(r, 4).value
            if m_v == month_label and np_v:
                row_map[str(np_v).strip()] = r

        for item in products:
            np_v = str(item.get('np', '')).strip()
            cant = int(item.get('cantidad_actual', 0))
            if np_v in row_map:
                target_row = row_map[np_v]
                ws.cell(target_row, 10).value = cant
                day_col_idx = 10 + day_num
                ws.cell(target_row, day_col_idx).value = cant
        wb.save(EXCEL_PATH)

def send_outlook(products, recipient):
    now = datetime.now()
    month_names = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    subject = f"INVENTARIO {now.day:02d} {month_names[now.month-1]} {now.year} TECNOLOGIA"
    
    items_by_section = {}
    for p in products:
        sec = p.get('seccion', 'VENTA')
        if sec not in items_by_section:
            items_by_section[sec] = []
        items_by_section[sec].append(p)

    html_parts = ["<html><body style='font-family: Calibri, Arial, sans-serif; color: #333333;'>"]
    for sec_title, prod_list in items_by_section.items():
        html_parts.append(f"<h2 style='color: #1F4E79; border-bottom: 2px solid #1F4E79; padding-bottom: 5px; margin-top: 25px;'>{sec_title}.</h2>")
        for item in prod_list:
            iva_str = " + IVA" if str(item.get('iva', '')).upper() == "SÍ" else ""
            cost_val = float(item.get('costo', 0))
            cost_formatted = f"${cost_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            cant = item.get('cantidad_actual', 0)
            
            html_parts.append("<table border='1' cellpadding='8' cellspacing='0' style='border-collapse: collapse; width: 100%; margin-bottom: 15px; border-color: #cccccc;'>")
            html_parts.append(f"  <tr><td style='background-color: #F2F4F8; font-weight: bold; width: 20%;'>N/P</td><td>{item.get('np')}</td></tr>")
            html_parts.append(f"  <tr><td style='background-color: #F2F4F8; font-weight: bold;'>Producto</td><td>{item.get('producto')}</td></tr>")
            html_parts.append(f"  <tr><td style='background-color: #F2F4F8; font-weight: bold;'>Costo</td><td><b>{cost_formatted}{iva_str}</b> / <span style='color: #D9534F; font-weight: bold;'>{cant} UNIDADES DISPONIBLES</span></td></tr>")
            html_parts.append(f"  <tr><td style='background-color: #F2F4F8; font-weight: bold;'>INGRESO</td><td>{item.get('fecha_ingreso', '')}</td></tr>")
            html_parts.append("</table>")

    html_parts.append("</body></html>")
    full_html = "\n".join(html_parts)

    import win32com.client as win32
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipient
    mail.Subject = subject
    mail.HTMLBody = full_html
    mail.Send()

if __name__ == '__main__':
    server = HTTPServer(('localhost', PORT), SyncHandler)
    print(f"🚀 Servidor Puente de Sincronización WebApp <-> Excel iniciado en http://localhost:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        server.server_close()
